import json
import requests
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich import box
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

console = Console()

PRICING_API = "https://prices.azure.com/api/retail/prices"

HEADERS = {
    "User-Agent": "azure-pricing-tool/1.0 (python-requests)",
    "Accept": "application/json",
}


def fetch_prices(filter_str: str, currency: str = "USD") -> list:
    items = []
    url = PRICING_API
    params = {"$filter": filter_str, "currencyCode": currency}
    while url:
        resp = requests.get(url, params=params, headers=HEADERS, timeout=30)
        if resp.status_code == 403:
            raise SystemExit(
                f"[HATA] Azure Pricing API erişimi engellendi (403 Forbidden).\n"
                f"URL: {resp.url}\n"
                f"Yanıt: {resp.text[:200]}\n\n"
                f"Bu genellikle ağ kısıtlamalarından kaynaklanır. "
                f"Scripti kendi makinenizde çalıştırın."
            )
        resp.raise_for_status()
        data = resp.json()
        items.extend(data.get("Items", []))
        next_page = data.get("NextPageLink")
        url = next_page
        params = {}
    return items


def find_price(items: list, price_type: str, term: str | None = None) -> float | None:
    for item in items:
        if item.get("type") != price_type:
            continue
        if term:
            if item.get("reservationTerm") == term:
                return item.get("retailPrice")
        else:
            return item.get("retailPrice")
    return None


# ── VM pricing ────────────────────────────────────────────────────────────────

def get_vm_prices(sku: str, region: str, currency: str, hours: int) -> dict:
    f = (
        f"armRegionName eq '{region}' and "
        f"armSkuName eq '{sku}' and "
        f"serviceName eq 'Virtual Machines'"
    )
    items = fetch_prices(f, currency)
    payg_hourly = find_price(items, "Consumption")
    res1yr_hourly = find_price(items, "Reservation", "1 Year")
    res3yr_hourly = find_price(items, "Reservation", "3 Years")

    def monthly(h):
        return round(h * hours, 4) if h is not None else None

    return {
        "name": sku,
        "payg": monthly(payg_hourly),
        "res1yr": monthly(res1yr_hourly),
        "res3yr": monthly(res3yr_hourly),
    }


# ── Disk pricing ──────────────────────────────────────────────────────────────

def get_disk_prices(disk_type: str, region: str, currency: str) -> dict:
    f = (
        f"armRegionName eq '{region}' and "
        f"skuName eq '{disk_type} LRS' and "
        f"serviceName eq 'Storage'"
    )
    items = fetch_prices(f, currency)
    payg = None
    for item in items:
        if item.get("type") == "Consumption" and "Disk" in item.get("productName", ""):
            payg = item.get("retailPrice")
            break
    return {"name": f"{disk_type} Managed Disk", "payg": payg, "res1yr": None, "res3yr": None}


# ── App Service pricing ───────────────────────────────────────────────────────

def get_app_service_prices(tier: str, region: str, currency: str, hours: int) -> dict:
    f = (
        f"armRegionName eq '{region}' and "
        f"skuName eq '{tier}' and "
        f"serviceName eq 'Azure App Service'"
    )
    items = fetch_prices(f, currency)
    payg_hourly = find_price(items, "Consumption")
    res1yr_hourly = find_price(items, "Reservation", "1 Year")
    res3yr_hourly = find_price(items, "Reservation", "3 Years")

    def monthly(h):
        return round(h * hours, 4) if h is not None else None

    return {
        "name": f"App Service {tier}",
        "payg": monthly(payg_hourly),
        "res1yr": monthly(res1yr_hourly),
        "res3yr": monthly(res3yr_hourly),
    }


# ── SQL Database pricing ──────────────────────────────────────────────────────

def get_sql_prices(tier: str, region: str, currency: str, hours: int) -> dict:
    f = (
        f"armRegionName eq '{region}' and "
        f"skuName eq '{tier}' and "
        f"serviceName eq 'SQL Database'"
    )
    items = fetch_prices(f, currency)
    payg_hourly = find_price(items, "Consumption")
    res1yr_hourly = find_price(items, "Reservation", "1 Year")
    res3yr_hourly = find_price(items, "Reservation", "3 Years")

    def monthly(h):
        return round(h * hours, 4) if h is not None else None

    return {
        "name": f"SQL {tier}",
        "payg": monthly(payg_hourly),
        "res1yr": monthly(res1yr_hourly),
        "res3yr": monthly(res3yr_hourly),
    }


# ── Storage Account pricing ───────────────────────────────────────────────────

def get_storage_prices(storage_type: str, gb: int, region: str, currency: str) -> dict:
    f = (
        f"armRegionName eq '{region}' and "
        f"skuName eq '{storage_type}' and "
        f"serviceName eq 'Storage' and "
        f"productName eq 'Blob Storage'"
    )
    items = fetch_prices(f, currency)
    price_per_gb = None
    for item in items:
        if item.get("type") == "Consumption":
            price_per_gb = item.get("retailPrice")
            break
    payg = round(price_per_gb * gb, 4) if price_per_gb is not None else None
    return {"name": f"Storage {storage_type} {gb}GB", "payg": payg, "res1yr": None, "res3yr": None}


# ── Helpers ───────────────────────────────────────────────────────────────────

def savings_pct(payg, reserved):
    if payg and reserved and payg > 0:
        return round((1 - reserved / payg) * 100, 1)
    return None

def fmt_price(val, count=1):
    if val is None:
        return "N/A"
    return f"${val * count:,.2f}"

def fmt_pct(val):
    if val is None:
        return "N/A"
    return f"{val}%"


# ── Terminal output ───────────────────────────────────────────────────────────

def print_section(title: str, rows: list, counts: list, color: str):
    table = Table(title=title, box=box.ROUNDED,
                  header_style=f"bold {color}", title_style=f"bold {color}", show_lines=True)
    table.add_column("Kaynak", style="white", min_width=28)
    table.add_column("Adet", justify="center", style="cyan")
    table.add_column("PAYG/ay", justify="right", style="yellow")
    table.add_column("1yr Rezervasyon/ay", justify="right", style="green")
    table.add_column("3yr Rezervasyon/ay", justify="right", style="bright_green")
    table.add_column("1yr Tasarruf%", justify="right", style="magenta")
    table.add_column("3yr Tasarruf%", justify="right", style="bright_magenta")

    total_payg = total_res1 = total_res3 = 0.0
    for row, count in zip(rows, counts):
        p, r1, r3 = row["payg"], row["res1yr"], row["res3yr"]
        s1, s3 = savings_pct(p, r1), savings_pct(p, r3)
        table.add_row(row["name"], str(count),
                      fmt_price(p, count), fmt_price(r1, count), fmt_price(r3, count),
                      fmt_pct(s1), fmt_pct(s3))
        if p: total_payg += p * count
        if r1: total_res1 += r1 * count
        if r3: total_res3 += r3 * count

    s1_t = savings_pct(total_payg, total_res1) if total_res1 else None
    s3_t = savings_pct(total_payg, total_res3) if total_res3 else None
    table.add_row("[bold]TOPLAM[/bold]", "",
                  f"[bold]${total_payg:,.2f}[/bold]",
                  f"[bold]${total_res1:,.2f}[/bold]" if total_res1 else "[bold]N/A[/bold]",
                  f"[bold]${total_res3:,.2f}[/bold]" if total_res3 else "[bold]N/A[/bold]",
                  fmt_pct(s1_t), fmt_pct(s3_t), style="bold")
    console.print(table)
    return total_payg, total_res1, total_res3


# ── Excel output ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
TOTAL_FILL  = PatternFill("solid", fgColor="D6E4F0")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FONT  = Font(bold=True, size=11)
CENTER = Alignment(horizontal="center", vertical="center")
thin   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row_idx, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER

def style_total(ws, row_idx, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill, cell.font, cell.alignment, cell.border = TOTAL_FILL, TOTAL_FONT, CENTER, BORDER

def write_sheet(wb, sheet_name: str, rows: list, counts: list):
    ws = wb.create_sheet(sheet_name)
    headers = ["Kaynak Adı", "Adet", "PAYG/ay ($)",
               "1yr Rezervasyon/ay ($)", "3yr Rezervasyon/ay ($)",
               "1yr Tasarruf%", "3yr Tasarruf%"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    totals = [0.0, 0.0, 0.0]
    for row, count in zip(rows, counts):
        p, r1, r3 = row["payg"], row["res1yr"], row["res3yr"]
        s1, s3 = savings_pct(p, r1), savings_pct(p, r3)
        ws.append([row["name"], count,
                   round(p * count, 2) if p else None,
                   round(r1 * count, 2) if r1 else None,
                   round(r3 * count, 2) if r3 else None,
                   s1, s3])
        if p: totals[0] += p * count
        if r1: totals[1] += r1 * count
        if r3: totals[2] += r3 * count

    s1_t = savings_pct(totals[0], totals[1]) if totals[1] else None
    s3_t = savings_pct(totals[0], totals[2]) if totals[2] else None
    total_row_idx = ws.max_row + 1
    ws.append(["TOPLAM", "",
               round(totals[0], 2) if totals[0] else None,
               round(totals[1], 2) if totals[1] else None,
               round(totals[2], 2) if totals[2] else None,
               s1_t, s3_t])
    style_total(ws, total_row_idx, len(headers))

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 4, 16)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if not cell.fill.fgColor.rgb or cell.fill.fgColor.rgb == "00000000":
                cell.border = BORDER
            cell.alignment = CENTER

    return totals[0], totals[1], totals[2]

def write_summary_sheet(wb, summary_data: list):
    ws = wb.create_sheet("Summary", 0)
    headers = ["Servis", "PAYG Toplam/ay ($)",
               "1yr Rezervasyon Toplam/ay ($)", "3yr Rezervasyon Toplam/ay ($)",
               "1yr Tasarruf%", "3yr Tasarruf%"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    grand = [0.0, 0.0, 0.0]
    for name, p, r1, r3 in summary_data:
        s1, s3 = savings_pct(p, r1), savings_pct(p, r3)
        ws.append([name,
                   round(p, 2) if p else None,
                   round(r1, 2) if r1 else None,
                   round(r3, 2) if r3 else None,
                   s1, s3])
        grand[0] += p or 0; grand[1] += r1 or 0; grand[2] += r3 or 0

    s1_g = savings_pct(grand[0], grand[1]) if grand[1] else None
    s3_g = savings_pct(grand[0], grand[2]) if grand[2] else None
    total_row_idx = ws.max_row + 1
    ws.append(["GENEL TOPLAM",
               round(grand[0], 2), round(grand[1], 2), round(grand[2], 2),
               s1_g, s3_g])
    style_total(ws, total_row_idx, len(headers))

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 4, 20)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = CENTER


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    with open("config.json") as f:
        cfg = json.load(f)

    region    = cfg["region"]
    currency  = cfg["currency"]
    hours     = cfg["hours_per_month"]
    resources = cfg["resources"]

    console.print(Panel(
        f"[bold cyan]Azure Fiyatlandırma Aracı[/bold cyan]\n"
        f"Bölge: [yellow]{region}[/yellow]  |  Para Birimi: [yellow]{currency}[/yellow]  |  Aylık saat: [yellow]{hours}[/yellow]",
        expand=False,
    ))

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    summary_data = []

    if resources.get("vms"):
        console.print("\n[bold cyan]VM fiyatları çekiliyor...[/bold cyan]")
        vm_rows, vm_counts = [], []
        for vm in resources["vms"]:
            console.print(f"  → {vm['sku']}")
            vm_rows.append(get_vm_prices(vm["sku"], region, currency, hours))
            vm_counts.append(vm["count"])
        p, r1, r3 = print_section("Virtual Machines", vm_rows, vm_counts, "cyan")
        write_sheet(wb, "VMs", vm_rows, vm_counts)
        summary_data.append(("Virtual Machines", p, r1, r3))

    if resources.get("disks"):
        console.print("\n[bold yellow]Disk fiyatları çekiliyor...[/bold yellow]")
        disk_rows, disk_counts = [], []
        for disk in resources["disks"]:
            console.print(f"  → {disk['type']}")
            disk_rows.append(get_disk_prices(disk["type"], region, currency))
            disk_counts.append(disk["count"])
        p, r1, r3 = print_section("Managed Disks", disk_rows, disk_counts, "yellow")
        write_sheet(wb, "Disks", disk_rows, disk_counts)
        summary_data.append(("Managed Disks", p, r1, r3))

    if resources.get("app_services"):
        console.print("\n[bold green]App Service fiyatları çekiliyor...[/bold green]")
        as_rows, as_counts = [], []
        for svc in resources["app_services"]:
            console.print(f"  → {svc['tier']}")
            as_rows.append(get_app_service_prices(svc["tier"], region, currency, hours))
            as_counts.append(svc["count"])
        p, r1, r3 = print_section("App Services", as_rows, as_counts, "green")
        write_sheet(wb, "App Services", as_rows, as_counts)
        summary_data.append(("App Services", p, r1, r3))

    if resources.get("sql_databases"):
        console.print("\n[bold magenta]SQL Database fiyatları çekiliyor...[/bold magenta]")
        sql_rows, sql_counts = [], []
        for db in resources["sql_databases"]:
            console.print(f"  → {db['tier']}")
            sql_rows.append(get_sql_prices(db["tier"], region, currency, hours))
            sql_counts.append(db["count"])
        p, r1, r3 = print_section("SQL Databases", sql_rows, sql_counts, "magenta")
        write_sheet(wb, "SQL", sql_rows, sql_counts)
        summary_data.append(("SQL Databases", p, r1, r3))

    if resources.get("storage_accounts"):
        console.print("\n[bold blue]Storage fiyatları çekiliyor...[/bold blue]")
        st_rows, st_counts = [], []
        for sa in resources["storage_accounts"]:
            console.print(f"  → {sa['type']} {sa['gb']}GB")
            st_rows.append(get_storage_prices(sa["type"], sa["gb"], region, currency))
            st_counts.append(1)
        p, r1, r3 = print_section("Storage Accounts", st_rows, st_counts, "blue")
        write_sheet(wb, "Storage", st_rows, st_counts)
        summary_data.append(("Storage Accounts", p, r1, r3))

    write_summary_sheet(wb, summary_data)
    excel_file = "azure_pricing_report.xlsx"
    wb.save(excel_file)
    console.print(f"\n[bold green]✓ Excel raporu kaydedildi:[/bold green] {excel_file}\n")


if __name__ == "__main__":
    main()
